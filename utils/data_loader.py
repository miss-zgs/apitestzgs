"""
多格式用例数据加载器
统一接口加载 YAML / JSON / Excel(.xlsx) / CSV 格式的测试数据
"""
import csv
import json
import logging
import os
from typing import List

import yaml

logger = logging.getLogger(__name__)

# 支持的文件格式与对应加载函数的映射
_LOADER_MAP = {}


def _register(ext: str):
    """装饰器：注册文件扩展名对应的加载函数"""
    def decorator(func):
        _LOADER_MAP[ext] = func
        return func
    return decorator


# ---------- 各格式加载器 ----------

@_register(".yaml")
@_register(".yml")
def _load_yaml(file_path: str) -> List[dict]:
    with open(file_path, "r", encoding="utf-8") as file:
        data = yaml.safe_load(file)
    return _ensure_list(data, file_path)


@_register(".json")
def _load_json(file_path: str) -> List[dict]:
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
    return _ensure_list(data, file_path)


@_register(".xlsx")
@_register(".xls")
def _load_excel(file_path: str) -> List[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("读取 Excel 需要安装 openpyxl: pip install openpyxl")

    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        logger.warning("Excel 文件 %s 没有有效数据行", file_path)
        return []

    headers = [str(cell).strip() if cell else f"col_{i}" for i, cell in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        row_dict = {}
        for header, value in zip(headers, row):
            row_dict[header] = _convert_cell_value(value)
        result.append(row_dict)
    return result


@_register(".csv")
def _load_csv(file_path: str) -> List[dict]:
    with open(file_path, "r", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        return [dict(row) for row in reader]


# ---------- 公开接口 ----------

def load_test_data(file_path: str) -> List[dict]:
    """
    统一入口：根据文件扩展名自动选择加载器，返回 List[dict]

    :param file_path: 数据文件的绝对或相对路径
    :return: 用例数据列表
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"数据文件不存在: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    loader = _LOADER_MAP.get(ext)
    if loader is None:
        supported = ", ".join(sorted(_LOADER_MAP.keys()))
        raise ValueError(f"不支持的文件格式 '{ext}'，当前支持: {supported}")

    logger.info("加载测试数据: %s (格式: %s)", file_path, ext)
    data = loader(file_path)
    logger.info("成功加载 %d 条用例数据", len(data))
    return data


def load_from_directory(directory: str, pattern: str = None) -> List[dict]:
    """
    加载目录下所有支持格式的数据文件，可选按文件名关键词过滤

    :param directory: 目录路径
    :param pattern: 文件名包含的关键词（可选）
    :return: 合并后的用例数据列表
    """
    if not os.path.isdir(directory):
        raise FileNotFoundError(f"目录不存在: {directory}")

    all_data = []
    for filename in sorted(os.listdir(directory)):
        ext = os.path.splitext(filename)[1].lower()
        if ext not in _LOADER_MAP:
            continue
        if pattern and pattern not in filename:
            continue
        file_path = os.path.join(directory, filename)
        all_data.extend(load_test_data(file_path))
    return all_data


# ---------- 辅助函数 ----------

def _ensure_list(data, file_path: str) -> List[dict]:
    """确保返回值是 list[dict]"""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return [data]
    raise TypeError(f"文件 {file_path} 的数据格式不合法，需要 list 或 dict，实际: {type(data)}")


def _convert_cell_value(value):
    """Excel 单元格值转换：尝试将字符串解析为 dict/list"""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith(("{", "[")):
            try:
                return json.loads(stripped)
            except json.JSONDecodeError:
                pass
    return value
